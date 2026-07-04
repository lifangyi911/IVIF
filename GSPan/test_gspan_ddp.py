# 在 2 张显卡上运行
# torchrun --nproc_per_node=2 apnn_gf2_psxl_ddp.py --dataset_name qb --batch_size 4
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import torch.backends.cudnn as cudnn
from torch.utils.data import DataLoader
import random
import os
from os.path import join
import time
import sys
#import gdal, osr
# from osgeo import gdal, osr
from dataset_h5 import *
from loss import *
from functions import *
import pansharp_metrics as mtc
from datetime import datetime
from pytorchtools import EarlyStopping
import openpyxl  # 保存到Excel表
from tqdm import tqdm
import argparse
from modelv5 import STDSfea2gsV2B
# from model_fusionnet_gs import FusionNetgs
# from network_swinfusion_gs_embedv2 import SwinFusion
import h5py
# from hqnr_torch import D_lambda_k, D_s_new
from hqnr_torch_fast import EfficientHQNR
import torch.nn.functional as F
import scipy.io as sio
from torch.utils.tensorboard import SummaryWriter
from wald_utilities import wald_protocol_v1, wald_protocol_v2
from utils.gaussian_splatting import generate_2D_gaussian_splatting_step, generate_2D_gaussian_splatting_batch
import torch.distributed as dist
from torch.utils.data.distributed import DistributedSampler
from torch.nn.parallel import DistributedDataParallel as DDP

parser = argparse.ArgumentParser(description='Args for FusionNet pansharpening')
parser.add_argument('-seed', type=int, default=2, help='seed')
parser.add_argument('--output_nc', type=int, default=8, help='output image channels')
parser.add_argument('--batch_size', type=int, default=16, help='batch_size') #default 16 for training, 1 for testing
parser.add_argument('--num_epochs', type=int, default=400, help='training epochs')
parser.add_argument('--lr', type=float, default=0.0004, help='output image channels')
# parser.add_argument('--step_size', type=int, default=100, help='每step_size个epoch更新一次学习率,变为原来的decay_rate倍')
# parser.add_argument('--decay_rate', type=float, default=0.1, help='每step_size个epoch更新一次学习率,变为原来的decay_rate倍')
parser.add_argument('--log_freq', type=int, default=10, help='screen output')
parser.add_argument('--patience', type=int, default=250, help='early stopping patience')
parser.add_argument('--dataset_name', type=str, default='wv3_4k', help='读取h5文件')
parser.add_argument('--gpuid', type=str, default='0,1,2,3', help='screen output')
parser.add_argument('--version', type=str, default=r'1',help='method version')
parser.add_argument('--resume', type=str, default='', help='path to checkpoint to resume from (e.g. .../net_ckpt_epoch200.pth)')
args = parser.parse_args()

import os
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "expandable_segments:True"
############## arguments
sensor = 'WV3'
if sensor=='GF2':
    bit = 10
else:
    bit = 11
spectral_num = args.output_nc
img_range = 2 ** bit   # 1.0  # 2047.0
disk_name = 'data7a'  # QB nvme4a, GF2: data7a, WV4-bands: nvme4b



# method = 'GS_highlight'
method = 'GS_nores'
# method = 'GS_res'
dataset = r'/raid/dataset/data_lifangyi/data/pansharpen/%s' % args.dataset_name
model_dir = r'/raid/dataset/data_lifangyi/apnn'
# eval_ckpt_path = os.path.join(model_dir,'net_best_ckpt.pth')
# eval_ckpt_path = r'/raid/dataset/data_lifangyi/apnn/logs_qb_GS/structure144V7pth/net_ckpt_epoch120.pth'
# eval_ckpt_path = r'/raid/dataset/data_lifangyi/apnn/logs_wv3_GS/3.27/net_ckpt_epoch300.pth'
# eval_ckpt_path = r'/raid/dataset/data_lifangyi/apnn/logs_%s_%s/%s_ckpt_epoch210.pth' % (args.dataset_name, method,method)
eval_ckpt_path = r'/raid/dataset/data_lifangyi/apnn/logs_gf2_GS/continuenolimit/net_ckpt_epoch300.pth'
# eval_ckpt_path = r'/raid/dataset/data_lifangyi/apnn/net_ckpt_epoch80.pth'
test_sample_dir = r'/raid/dataset/data_lifangyi/apnn/results/%s/%s_results/test_sample_v%s' % (args.dataset_name,method,args.version)
origin_test_sample_dir = r'/raid/dataset/data_lifangyi/apnn/results/%s/%s_results/origin_test_sample_v%s' % (args.dataset_name,method,args.version)
record_dir = r'/raid/dataset/data_lifangyi/apnn/results/record_v%s' % (args.version)
logs_dir = r"./logs_%s_%s/" % (args.dataset_name, method)

if not os.path.exists(model_dir):
    os.makedirs(model_dir)
if not os.path.exists(test_sample_dir):
    os.makedirs(test_sample_dir)
if not os.path.exists(origin_test_sample_dir):
    os.makedirs(origin_test_sample_dir)
if not os.path.exists(record_dir):
    os.makedirs(record_dir)
if not os.path.exists(logs_dir):
    os.makedirs(logs_dir)


print(f'Method: {method}, batch_size: {args.batch_size}, Version: {args.version}')


mode = "a" if args.resume else "w"
f = open("./logs_%s_%s/log_v%s.txt" % (args.dataset_name, method, args.version), 'a')


## Device configuration
os.environ['CUDA_DEVICE_ORDER'] = 'PCI_BUS_ID'
os.environ['CUDA_VISIBLE_DEVICES'] = args.gpuid
# device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
# dist.init_process_group(backend='nccl')
# local_rank = int(os.environ["LOCAL_RANK"])
# torch.cuda.set_device(local_rank)
# device = torch.device("cuda", local_rank)
# print(f'Using device: {device}')
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
local_rank = 0
print(f'Using device: {device}')
if args.resume and local_rank == 0:
    f.write(f"\n\n{'=' * 20} RESUME TRAINING AT {datetime.now()} {'=' * 20}\n")

# 初始化损失函数
FastHQNR = EfficientHQNR(sensor=sensor, ratio=4, channels=4, device='cuda')

############## Loading datasets
print('===> Loading datasets')

# train_low_set = Dataset_Pro(os.path.join(dataset, 'train_%s.h5' % args.dataset_name),img_scale=img_range,augment=True)
# train_sampler = DistributedSampler(train_low_set)
# train_loader = DataLoader(train_low_set, batch_size=args.batch_size, sampler=train_sampler, num_workers=2, pin_memory=False, drop_last=True)
# print('len(train_loader)', len(train_loader))
#
# valid_low_set = Dataset_Pro(os.path.join(dataset, 'valid_%s.h5' % args.dataset_name),img_scale=img_range)
# valid_loader = DataLoader(valid_low_set, batch_size=args.batch_size, num_workers=2, shuffle=False, pin_memory=False, drop_last=True)

test_low_set = MultiExmTest_h5(os.path.join(dataset, 'test_%s_multiExm1.h5' % args.dataset_name),img_scale=img_range)
test_low_loader = DataLoader(test_low_set, batch_size=1, num_workers=1, shuffle=False, pin_memory=False, drop_last=False)

# test_full_set = MultiExmTest_h5(os.path.join(dataset, 'TestData_%s_fr_patches512_float64.h5' % args.dataset_name),img_scale=img_range)
# test_full_set = MultiExmTest_h5(os.path.join(dataset, 'TestData_%s_fr_patches1024_nolap.h5' % args.dataset_name),img_scale=img_range)
# test_full_set = MultiExmTest_h5(os.path.join(dataset, 'TestData_%s_fr.h5' % args.dataset_name),img_scale=img_range)
# test_full_loader = DataLoader(test_full_set, batch_size=1, num_workers=1, shuffle=False, pin_memory=False, drop_last=False)
test_full_set = MultiExmTest_h5(os.path.join(dataset, 'test_%s_OrigScale_multiExm1.h5'  % args.dataset_name),img_scale=img_range)
test_full_loader = DataLoader(test_full_set, batch_size=1, num_workers=1, shuffle=False, pin_memory=False, drop_last=False)

criterionL1 = nn.L1Loss().to(device)
criterionSmoothL1 = nn.SmoothL1Loss(beta = 0.5).to(device)
criterionMSE = nn.MSELoss().to(device)

# net = model = STDSfea2gsV2B(spectral_num=spectral_num).to(device)
# net = model = STDSfea2gsV2B(spectral_num=spectral_num, inchannel=32, channel=48).to(device)
net = model = STDSfea2gsV2B(spectral_num=spectral_num,sensor=sensor,shuffle_scale1 = 2, shuffle_scale2 = 1).to(device)

optimizer = torch.optim.Adam(net.parameters(), lr=args.lr, betas=(0.9, 0.99))

# 1. 设定参数
warmup_epochs = 5      # 预热代数
T_max = args.num_epochs - warmup_epochs  # 余弦退火的总周期（总代数减去预热代数）
eta_min = 1e-7         # 最小学习率，通常设为一个很小的值

# 2. 定义 Warmup 调度器：从 0.1 * lr 线性增加到 lr
scheduler_warmup = torch.optim.lr_scheduler.LinearLR(
    optimizer,
    start_factor=0.1,
    total_iters=warmup_epochs
)

# 3. 定义余弦退火调度器
# T_max 是到达 eta_min 所需的迭代次数
scheduler_main = torch.optim.lr_scheduler.CosineAnnealingLR(
    optimizer,
    T_max=T_max,
    eta_min=eta_min
)

# 4. 组合调度器
# milestones=[warmup_epochs] 表示在第 5 个 epoch 结束后切换到余弦退火
scheduler = torch.optim.lr_scheduler.SequentialLR(
    optimizer,
    schedulers=[scheduler_warmup, scheduler_main],
    milestones=[warmup_epochs]
)

# 2. 恢复训练逻辑
start_epoch = 1
if args.resume:
    if os.path.isfile(args.resume):
        if local_rank == 0:
            print(f"==> Resuming from checkpoint: {args.resume}")

        # 使用 map_location 确保加载到正确的 GPU
        checkpoint = torch.load(args.resume, map_location=device)

        # 加载权重 (处理 DDP 保存时可能带有的 'module.' 前缀)
        state_dict = checkpoint['net']
        from collections import OrderedDict

        new_state_dict = OrderedDict()
        for k, v in state_dict.items():
            name = k[7:] if k.startswith('module.') else k  # 去掉 module. 前缀
            new_state_dict[name] = v
        net.load_state_dict(new_state_dict)

        # # 加载优化器和调度器状态
        # optimizer.load_state_dict(checkpoint['optimizer'])
        # scheduler.load_state_dict(checkpoint['scheduler'])
        # 如果更换损失函数，重置优化器和调度器
        start_epoch = checkpoint['epoch'] + 1
        optimizer = torch.optim.Adam(net.parameters(), lr=0.75e-4, betas=(0.9, 0.99))
        scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
            optimizer,
            T_max=T_max - start_epoch,
            eta_min=eta_min
        )
        # for _ in range(start_epoch - 1):
        #     scheduler.step()
        # if local_rank == 0:
        #     print(f"==> Loaded checkpoint successfully, starting from epoch {start_epoch}")
    else:
        if local_rank == 0:
            print(f"==> No checkpoint found at: {args.resume}")
if local_rank == 0:
    print(args)
    print('')
    print(optimizer)
    print('')
# net = DDP(net, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)


def load_checkpoint(model, checkpoint_path):
    """通用权重加载函数，处理 DDP 的 module. 前缀"""
    checkpoint = torch.load(checkpoint_path, map_location=device)
    # 如果是完整的 checkpoint 字典
    if 'net' in checkpoint:
        state_dict = checkpoint['net']
    else:
        state_dict = checkpoint

    from collections import OrderedDict
    new_state_dict = OrderedDict()
    for k, v in state_dict.items():
        # 如果权重里有 module. 而当前模型没有，或者反之，进行处理
        name = k[7:] if k.startswith('module.') else k
        new_state_dict[name] = v

    # 如果 model 本身是 DDP 包装的，需要加载到 model.module
    if isinstance(model, torch.nn.parallel.DistributedDataParallel):
        model.module.load_state_dict(new_state_dict)
    else:
        model.load_state_dict(new_state_dict)
    print(f"Successfully loaded weights from {checkpoint_path}")

def net_low_eval(data_loader, model, eval_ckpt_path, f):
    load_checkpoint(model, eval_ckpt_path)
    model.eval()
    print("len(data_loader)", len(data_loader))

    ERGAS, SAM, Q2n, SCC = [], [], [], []
    with torch.no_grad():
        for index, batch in enumerate(data_loader,1):
            input_pan = batch['pan'].to(device)
            input_lr_u = batch['lms'].to(device)
            input_lr = batch['ms'].to(device)
            target = batch['gt'].to(device)

            batch_gs_parameters = net(input_lr_u, input_pan)
            gt_size = target.shape[2:]  # [h_gt, w_gt]

            psh = generate_2D_gaussian_splatting_batch(sr_size=gt_size, gs_parameters=batch_gs_parameters,
                                                       scale=1,
                                                       sample_coords=None,
                                                       scale_modify=[1, 1],
                                                       default_step_size=1.2,
                                                       cuda_rendering=True,
                                                       mode='scale',
                                                       if_dmax=True,
                                                       spectral_num=spectral_num,
                                                       dmax_mode='fix',
                                                       dmax=0.5)


            # print(f'psh before: {psh.mean()}, pan shape: {psh.shape}')

            psh = psh + input_lr_u

            # visualization the position of Gaussian
            # psh = psh + input_lr_u*0.002

            # visualization the highlight of Gaussian
            # psh = psh + input_lr_u

            psh = psh * img_range

            psh = trim_image(psh, L=0, R=2 ** bit-1)

            target, psh = target.detach().cpu().numpy(), psh.detach().cpu().numpy()
            target, psh = np.transpose(target, (0, 2, 3, 1)), np.transpose(psh, (0, 2, 3, 1))
            target, psh = target[0], psh[0]

            ERGAS.append(mtc.ERGAS(target, psh, 4))
            SAM.append(mtc.SAM(target, psh, eps=0.0005))  # , eps=0.0005
            Q2n.append(mtc.Q2n(target, psh, q_block_size=32, q_shift=32))
            SCC.append(mtc.SCC(target, psh))

            ## 将 每个 Fused RR image 保存为 一个 mat 文件
            filename_mat = '%s/fused_rr%d.mat' % (test_sample_dir, index)
            sio.savemat(filename_mat, {'fus': psh})

    print("SAM, SCC, ERGAS, Q2n")
    print("%.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf " %
          (np.mean(SAM), np.std(SAM), np.mean(SCC), np.std(SCC), np.mean(ERGAS), np.std(ERGAS), np.mean(Q2n), np.std(Q2n)))
    print(Q2n)

    print("SAM, SCC, ERGAS, Q2n", file=f, flush=True)
    print("%.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf " %
          (np.mean(SAM), np.std(SAM), np.mean(SCC), np.std(SCC), np.mean(ERGAS), np.std(ERGAS), np.mean(Q2n), np.std(Q2n)), file=f, flush=True)

    wb_low_eval_metric = openpyxl.Workbook()
    ws_low_eval_metric = wb_low_eval_metric.create_sheet('sheet1',0)

    metrics_name_list = ["SAM", "SCC", "ERGAS", "Q2n"]
    metrics_list = ["%.4lf±%.4lf" % (np.mean(SAM), np.std(SAM)) , "%.4lf±%.4lf" % (np.mean(SCC), np.std(SCC)), "%.4lf±%.4lf" % (np.mean(ERGAS), np.std(ERGAS)), "%.4lf±%.4lf" % (np.mean(Q2n), np.std(Q2n))]

    for i in range(len(metrics_name_list)):
        ws_low_eval_metric.cell(row=1, column=i + 1).value = metrics_name_list[i]
        ws_low_eval_metric.cell(row=2, column=i + 1).value = metrics_list[i]

    wb_low_eval_metric.save('%s/low_eval_metric_record.xlsx' % record_dir)


def split_and_joint_pansharpen(pan, lms, split_size, overlap_size,
                               model, img_range,
                               default_step_size=1.2, dmax=1.0):
    """
    pan: [1, 1, H, W] 原始全色图
    lms: [1, 4, H, W] 已上采样的多光谱图
    split_size: 分块大小 (如 64)
    overlap_size: 重叠大小 (如 8)
    """
    import math
    b, _, h_raw, w_raw = pan.shape
    device = pan.device

    # 1. 计算分块数量和补齐(Padding)
    stride = split_size - overlap_size
    tile_nums_h = math.ceil((h_raw - overlap_size) / stride)
    tile_nums_w = math.ceil((w_raw - overlap_size) / stride)

    pad_h = tile_nums_h * stride + overlap_size - h_raw
    pad_w = tile_nums_w * stride + overlap_size - w_raw

    # 对两个输入同时进行 Padding
    pan_pad = F.pad(pan, (0, pad_w, 0, pad_h), mode='reflect')
    lms_pad = F.pad(lms, (0, pad_w, 0, pad_h), mode='reflect')

    # 2. 创建输出画布和权重累加图 (用于处理重叠区域的平滑)
    output_canvas = torch.zeros_like(lms_pad)
    count_map = torch.zeros_like(lms_pad)

    # 创建余弦或线性权重窗，消除接缝（比你原代码的 crop 逻辑更平滑）
    window = torch.ones((1, 1, split_size, split_size), device=device)
    # 这里可以简单的给边缘加权，防止硬拼接痕迹
    for i in range(overlap_size):
        v = (i + 1) / overlap_size
        window[:, :, i, :] *= v
        window[:, :, -(i + 1), :] *= v
        window[:, :, :, i] *= v
        window[:, :, :, -(i + 1)] *= v

    # 3. 滑动窗口推理
    for h_idx in range(tile_nums_h):
        for w_idx in range(tile_nums_w):
            h_start = h_idx * stride
            w_start = w_idx * stride
            h_end = h_start + split_size
            w_end = w_start + split_size

            # 同步裁剪两个输入
            tile_pan = pan_pad[:, :, h_start:h_end, w_start:w_end]
            tile_lms = lms_pad[:, :, h_start:h_end, w_start:w_end]

            # 模型推理
            with torch.no_grad():
                # 注意：这里调用你的 modelv4 结构
                # 假设 scale_modify 在这里设为固定的 [1,1] 因为已经在高分辨率空间
                batch_gs_params = net(tile_lms, tile_pan)

                # 渲染当前小块
                tile_out = generate_2D_gaussian_splatting_batch(
                    sr_size=(split_size, split_size),
                    gs_parameters=batch_gs_params,
                    scale=1,  # 分块推理时，scale 始终为 1
                    sample_coords=None,
                    scale_modify=[1, 1],
                    default_step_size=1.2,
                    cuda_rendering=True,
                    mode='scale',
                    if_dmax=True,
                    dmax_mode='fix',
                    dmax=0.1
                )

                tile_out = tile_out*0.65 + tile_lms

            # 累加到输出画布
            output_canvas[:, :, h_start:h_end, w_start:w_end] += tile_out * window
            count_map[:, :, h_start:h_end, w_start:w_end] += window

    # 4. 除以权重图，得到平均值，并裁掉最外圈的 Padding
    final_output = output_canvas / (count_map + 1e-8)
    return final_output[:, :, :h_raw, :w_raw]

def net_full_eval(data_loader, model, eval_ckpt_path, f):
    load_checkpoint(model, eval_ckpt_path)
    model.eval()

    HQNR_set, D_lambda_k_set, D_s_set = [], [], []
    print("len(data_loader)", len(data_loader))

    with torch.no_grad():
        for index, batch in enumerate(data_loader):
            input_pan = batch['pan'].to(device)
            input_lr = batch['ms'].to(device)
            input_lr_u = batch['lms'].to(device)
            # input_pan_l = F.interpolate(input_pan, scale_factor=0.25, mode='nearest')
            input_pan_l = wald_protocol_v2(input_lr, input_pan, ratio=4, sensor=sensor, channels=args.output_nc)  # mode='nearest'
            print(f'input_pan_l: {input_pan_l.shape} input_pan: {input_pan.shape} input_lr_u: {input_lr_u.shape}')
            batch_gs_parameters = net(input_lr_u, input_pan)
            gt_size = input_pan.shape[2:]  # [h_gt, w_gt]

            psh = generate_2D_gaussian_splatting_batch(sr_size=gt_size, gs_parameters=batch_gs_parameters,
                                                       scale=1,
                                                       sample_coords=None,
                                                       scale_modify=[1, 1],
                                                       default_step_size=1.2,
                                                       cuda_rendering=True,
                                                       mode='scale',
                                                       if_dmax=True,
                                                       spectral_num=spectral_num,
                                                       dmax_mode='fix',
                                                       dmax=0.1)
            # psh = psh - F.avg_pool2d(psh, kernel_size=5, stride=1, padding=2)

            output = psh + input_lr_u
            # output = psh + input_lr_u*0.0002
            # output = input_lr_u

            output = trim_image(output, L=0, R=1)
      
            
            # ## new and fast version
            FastHQNR_value,D_lambda_k_value,D_s_value = FastHQNR(output, input_lr_u, input_lr, input_pan, input_pan_l)
            HQNR_value = FastHQNR_value

            HQNR_set.append(HQNR_value.item()), D_lambda_k_set.append(D_lambda_k_value.item()), D_s_set.append(D_s_value.item())
            print('Test %d' % index, 'D_spe',D_lambda_k_value.item(),'D_spa',D_s_value.item(),'HQNR',HQNR_value.item())

            ##  将 每个 Fused FR image 保存为 一个 mat 文件
            output = output.detach().cpu().numpy()
            output = np.transpose(output, (0, 2, 3, 1))
            print('Saving... %d' % (index+1))
            filename_mat = '%s/fused_fr%d.mat' % (origin_test_sample_dir, index+1)
            sio.savemat(filename_mat, {'fus': output[0]})

            # 清理缓存
            import gc
            torch.cuda.empty_cache()
            gc.collect()

    print("D_lambda_k, D_s, HQNR")
    print("%.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf" %
          (np.mean(D_lambda_k_set), np.std(D_lambda_k_set), np.mean(D_s_set), np.std(D_s_set), np.mean(HQNR_set), np.std(HQNR_set)))
    print("D_lambda_k, D_s, HQNR", file=f, flush=True)
    print("%.4lf±%.4lf, %.4lf±%.4lf, %.4lf±%.4lf" %
          (np.mean(D_lambda_k_set), np.std(D_lambda_k_set), np.mean(D_s_set), np.std(D_s_set), np.mean(HQNR_set), np.std(HQNR_set)), file=f, flush=True)

    wb_full_eval_metric = openpyxl.Workbook()
    ws_full_eval_metric = wb_full_eval_metric.create_sheet('sheet1',0)

    metrics_name_list = ["D_lambda_k", "D_s", "HQNR"]
    metrics_list = ["%.4lf±%.4lf" % (np.mean(D_lambda_k_set), np.std(D_lambda_k_set)),
                    "%.4lf±%.4lf" % (np.mean(D_s_set), np.std(D_s_set)), "%.4lf±%.4lf" % (np.mean(HQNR_set), np.std(HQNR_set))]

    for i in range(len(metrics_name_list)):
        ws_full_eval_metric.cell(row=1, column=i + 1).value = metrics_name_list[i]
        ws_full_eval_metric.cell(row=2, column=i + 1).value = metrics_list[i]

    wb_full_eval_metric.save('%s/full_eval_metric_record.xlsx' % record_dir)

if "__main__" == __name__:
    print('Begin Time: ', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    # Train

    # train(train_loader, test_low_loader, args.num_epochs, args.patience, f)
    # print('==>Train 时长： {:.2f}h\n'.format((time.time() - start) / 3600))
    # print('==>Train 时长： {:.2f}h\n'.format((time.time() - start) / 3600), file=f, flush=True)
    ## Eval.
    print('Evaluating...')
    net_low_eval(test_low_loader, net, eval_ckpt_path, f)
    start = time.time()
    net_full_eval(test_full_loader, net, eval_ckpt_path, f)

    print('==>总时长： {:.2f}s\n'.format((time.time() - start)), file=f, flush=True)
    print('End Time: ', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # ========== 1. 模型参数量统计 ==========
    total_params = sum(p.numel() for p in net.parameters())
    trainable_params = sum(p.numel() for p in net.parameters() if p.requires_grad)
    print(f"\n{'=' * 50}")
    print(f"Model Statistics")
    print(f"{'=' * 50}")
    print(f"Total parameters: {total_params / 1e6:.2f}M ({total_params:,})")
    print(f"Trainable parameters: {trainable_params / 1e6:.2f}M ({trainable_params:,})")
    print(f"Non-trainable: {(total_params - trainable_params) / 1e6:.2f}M")

    # # ========== 2. FLOPs统计 ==========
    # print(f"\n{'=' * 50}")
    # print(f"FLOPs Statistics")
    # print(f"{'=' * 50}")
    # net.eval()
    # with torch.no_grad():
    #     batch = next(iter(test_low_loader))
    #     input_pan = batch['pan'].to(device)
    #     input_lr_u = batch['lms'].to(device)
    #
    #     # 计算FLOPs
    #     flops, params = profile(net, inputs=(input_lr_u, input_pan), verbose=False)
    #     flops, params = clever_format([flops, params], "%.3f")
    #     print(f"FLOPs: {flops}")
    #     print(f"Params (thop): {params}")
